
import os
import datetime
from openpyxl import load_workbook, Workbook
from vendor.handreamnet import handreamnet_switch
from vendor.aruba import aruba_switch

"""
향후 추가 기능
1. Version 확인
 - 각 모델별 명령어 출력 결과 필요
2. 월별 log 분석 후 로그인 성공 및 실패 횟수 파싱
 - 각 모델별 로그 명령어 및 로그 출력결과 필요
3. 아루바 스위치 추가
 - 아루바 스위치 "no page" 명령어 확인 필요
 - clock 관련 명령어 확인 필요
"""

def read_connection_info_from_file(filename="device_connection_infos.txt"):
    current_directory = os.getcwd()
    file_path = os.path.join(current_directory, filename)
    connection_infos = []

    try:
        with open(file_path, 'r') as file:
            for line in file:
                device_type, ip, port, username, password, enable_password = line.split(",")
                details = {
                    'device_type': device_type, 
                    'ip': ip,
                    'port': int(port),
                    'username': username,
                    'password': password,
                    'secret': enable_password,
                    'session_log': f'session_logs/{ip}_session_log.txt'
                }
                connection_infos.append(details)
    except FileNotFoundError:
        print(f"파일을 찾을 수 없음: {file_path}")
    except Exception as e:
        print(f"파일 읽기 오류: {e}")

    return connection_infos

def main():
    connection_infos = read_connection_info_from_file()

    current_directory = os.getcwd()
    backup_directory = os.path.join(current_directory, "backup")
    session_log_directory = os.path.join(current_directory, "session_logs")
    excel_file_path = os.path.join(current_directory, 'device_info.xlsx')

    if os.path.exists(backup_directory):
        print("동일한 이름의 폴더가 이미 존재합니다. 해당 폴더에 설정 파일을 백업합니다.")
    else:
        os.mkdir(backup_directory)

    if os.path.exists(session_log_directory):
        print("동일한 이름의 폴더가 이미 존재합니다. 해당 폴더에 세션 로그를 저장합니다.")
    else:
        os.mkdir(session_log_directory)

    if os.path.exists(excel_file_path):
        print("동일한 이름의 엑셀 파일이 이미 존재합니다. 해당 엑셀 파일에 데이터를 추가합니다.")
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["IP", "Hostname", "Fan Status", "Temperature", "Uptime", "CPU Usage", "Memory Usage"])

    for connection_info in connection_infos:
        current_time = datetime.datetime.now()
        # 한드림넷
        if 'handreamnet' in connection_info['device_type']:
            connection_info['device_type'] = 'generic' if connection_info['device_type'] == 'handreamnet' else connection_info['device_type']
            connection = handreamnet_switch(connection_info).connect_to_switch()
            if connection is None:
                print(f"{connection_info['ip']}: 연결 실패")
                continue
            formatted_time = current_time.strftime("clock %H:%M:%S %d %m %Y")
            connection.send_command_timing(formatted_time)
            print(f"{connection_info['ip']}: 명령어 입력 - {formatted_time}")

            running_config = connection.send_command_timing('show running-config')
            print(f"{connection_info['ip']}: 명령어 입력 - show running-config")

            hostname = handreamnet_switch(connection_info).hostname(running_config)
            print(f"{connection_info['ip']}: 호스트네임 - {hostname}")

            file_path = os.path.join(backup_directory, f"{connection_info['ip']}_{hostname}.txt")
            with open(file_path, "w") as file:
                file.write(running_config)
                print(f"{connection_info['ip']}: running-config 저장 완료")

            fan_status = connection.send_command_timing('show system fan')
            print(f"{connection_info['ip']}: 명령어 입력 - show system fan" + '\n' + fan_status)

            temperature = connection.send_command_timing('show system temperature')
            print(f"{connection_info['ip']}: 명령어 입력 - show system temperature" + '\n' + temperature)

            uptime = connection.send_command_timing('show system uptime')
            print(f"{connection_info['ip']}: 명령어 입력 - show system uptime" + '\n' + uptime)

            cpu_usage = connection.send_command_timing('show system cpu-load')
            print(f"{connection_info['ip']}: 명령어 입력 - show system cpu-load" + '\n' + cpu_usage)

            memory_usage = connection.send_command_timing('show system memory')
            print(f"{connection_info['ip']}: 명령어 입력 - show system memory" + '\n' + memory_usage)
            
            # 파싱
            fan_status = handreamnet_switch(connection_info).fan_status(fan_status)
            temperature = handreamnet_switch(connection_info).temperature(temperature)
            uptime = handreamnet_switch(connection_info).uptime(uptime)
            cpu_usage = handreamnet_switch(connection_info).cpu_usage(cpu_usage)
            memory_usage = handreamnet_switch(connection_info).memory_usage(memory_usage)

            connection.disconnect()

            ws.append([connection_info['ip'], hostname, fan_status, temperature, uptime, cpu_usage, memory_usage])

            while True:
                try:
                    wb.save(excel_file_path)
                    break
                except PermissionError:
                    print("엑셀 파일이 다른 프로그램에서 열려 있습니다. 파일을 닫고 Enter 키를 눌러 계속하세요.")
                    input()
        # 아루바
        elif 'aruba_os' in connection_info['device_type']:
            connection = aruba_switch(connection_info).connect_to_switch()
            if connection is None:
                print(f"{connection_info['ip']}: 연결 실패")
                continue
            formatted_time = current_time.strftime("clock %H:%M:%S %d %m %Y") # 수정필요
            connection.send_command_timing(formatted_time)
            print(f"{connection_info['ip']}: 명령어 입력 - {formatted_time}")

            running_config = connection.send_command_timing('show running-config')
            print(f"{connection_info['ip']}: 명령어 입력 - show running-config")

            hostname = aruba_switch(connection_info).hostname(running_config)
            print(f"{connection_info['ip']}: 호스트네임 - {hostname}")

            file_path = os.path.join(backup_directory, f"{connection_info['ip']}_{hostname}.txt")
            with open(file_path, "w") as file:
                file.write(running_config)
                print(f"{connection_info['ip']}: running-config 저장 완료")

            information = connection.send_command_timing('show system information')
            print(f"{connection_info['ip']}: 명령어 입력 - show system information" + '\n' + information)

            fan_status = connection.send_command_timing('show system fan')
            print(f"{connection_info['ip']}: 명령어 입력 - show system fan" + '\n' + fan_status)

            temperature = connection.send_command_timing('show system temperature')
            print(f"{connection_info['ip']}: 명령어 입력 - show system temperature" + '\n' + temperature)

            # 파싱
            fan_status = aruba_switch(connection_info).fan_status(fan_status)
            temperature = aruba_switch(connection_info).temperature(temperature)
            uptime = aruba_switch(connection_info).uptime(information)
            cpu_usage = aruba_switch(connection_info).cpu_usage(information)
            memory_usage = aruba_switch(connection_info).memory_usage(information)

            connection.disconnect()

            ws.append([connection_info['ip'], hostname, fan_status, temperature, uptime, cpu_usage, memory_usage])

            while True:
                try:
                    wb.save(excel_file_path)
                    break
                except PermissionError:
                    print("엑셀 파일이 다른 프로그램에서 열려 있습니다. 파일을 닫고 Enter 키를 눌러 계속하세요.")
                    input()
        else:
            print(f"{connection_info['ip']}: 지원하지 않는 장비 타입")
if __name__ == "__main__":
    main()